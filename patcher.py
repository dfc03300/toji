import json
import sys
from copy import copy

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side

DEFAULT_FONT = Font(name="맑은 고딕", size=11)
thin = Side(style="thin", color="C9D4D1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

SHEET_NAME = "자동정리"


def coerce(raw):
    if raw is None or raw == "":
        return raw
    text = str(raw).replace(",", "").strip()
    if not text:
        return raw
    try:
        as_float = float(text)
        return int(as_float) if as_float == int(as_float) else as_float
    except ValueError:
        return raw


def apply(xlsx_path, patches):
    wb = openpyxl.load_workbook(xlsx_path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"'{SHEET_NAME}' 시트를 찾지 못했습니다.")
    ws = wb[SHEET_NAME]
    for patch in patches:
        row = int(patch["row"])
        col = int(patch["col"])
        val = coerce(patch.get("value"))
        cell = ws.cell(row, col, val)
        cell.font = copy(DEFAULT_FONT)
        cell.border = border
        cell.alignment = Alignment(vertical="center")
        fmt = patch.get("format")
        if fmt:
            cell.number_format = fmt
        elif isinstance(val, (int, float)):
            cell.number_format = "#,##0"
    wb.save(xlsx_path)


if __name__ == "__main__":
    patches = json.loads(sys.stdin.read())
    apply(sys.argv[1], patches)
    print("ok", flush=True)
