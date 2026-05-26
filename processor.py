import json
import os
import re
import sys
import threading
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


HEADERS = [
    "기호",
    "소재지",
    "지번",
    "시점",
    "면적",
    "지목",
    "용도지역",
    "이용상황",
    "도로교통",
    "형상",
    "지세",
    "목적/거래가격",
    "단가(원/㎡)",
    "개별공시지가",
    "개공비율",
    "검토",
    "시점수정치",
    "시점수정",
    "크롤링상태",
]

DEFAULT_FONT = Font(name="맑은 고딕", size=11)
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True)
CRAWLED_FONT = Font(name="맑은 고딕", size=11, color="C00000")

SIDO_CACHE = {}
SIGUNGU_CACHE = {}
DONGRI_CACHE = {}

MAX_CRAWL_WORKERS = 6
_log_lock = threading.Lock()


def public_base_url():
    return os.environ.get("PUBLIC_BASE_URL", "http://127.0.0.1:5180").rstrip("/")


def log(message):
    with _log_lock:
        print(message, flush=True)


def norm(value):
    return str(value or "").strip()


def to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = re.sub(r"[^0-9.\-]", "", str(value))
    if not text:
        return None
    try:
        return float(text) if "." in text else int(text)
    except ValueError:
        return None


def split_jibun(value):
    text = norm(value)
    san = "1"
    if text.startswith("산"):
        san = "2"
        text = text[1:].strip()
    main, sub = (text.split("-", 1) + ["0"])[:2] if "-" in text else (text, "0")
    main_digits = re.sub(r"\D", "", str(main))
    sub_digits = re.sub(r"\D", "", str(sub))
    if not main_digits:
        return san, "", ""
    return san, main_digits.zfill(4), (sub_digits or "0").zfill(4)


def display_jibun(value):
    text = norm(value)
    return int(text) if re.fullmatch(r"\d+", text) else text


def post_json(path, params):
    data = urllib.parse.urlencode(params).encode("utf-8")
    req = urllib.request.Request(
        f"https://www.realtyprice.kr{path}",
        data=data,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "X-Requested-With": "XMLHttpRequest",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=12) as resp:
        return json.loads(resp.read().decode("utf-8"))


def get_list(path, params):
    data = post_json(path, params)
    return data.get("model", {}).get("list", []) or []


def code_lookup(si_gun_gu, dong):
    parts = si_gun_gu.split()
    if len(parts) < 2:
        return None
    sido_name = parts[0]
    sigungu_name = parts[1]
    dong_name = dong.strip()

    if "sido" not in SIDO_CACHE:
        SIDO_CACHE["sido"] = get_list("/notice/m/bjd/getSido.do", {"notice_year": ""})
    def find_name(items, name):
        exact = next((x for x in items if x.get("NAME") == name), None)
        if exact:
            return exact
        return next((x for x in items if x.get("NAME", "").startswith(name) or name.startswith(x.get("NAME", "")[:2])), None)

    sido = find_name(SIDO_CACHE["sido"], sido_name)
    if not sido:
        return None

    sig_key = sido["CODE"]
    if sig_key not in SIGUNGU_CACHE:
        SIGUNGU_CACHE[sig_key] = get_list("/notice/m/bjd/getSigungu.do", {"notice_year": "", "reg1": sig_key})
    sigungu = find_name(SIGUNGU_CACHE[sig_key], sigungu_name)
    if not sigungu:
        return None

    dong_key = sigungu["CODE"]
    if dong_key not in DONGRI_CACHE:
        DONGRI_CACHE[dong_key] = get_list("/notice/m/bjd/getDongri.do", {"notice_year": "", "reg": dong_key})
    dongri = find_name(DONGRI_CACHE[dong_key], dong_name)
    if not dongri:
        return None

    return sido, sigungu, dongri


def crawl_price(si_gun_gu, dong, jibun, trade_date):
    codes = code_lookup(norm(si_gun_gu), norm(dong))
    if not codes:
        return None, "주소코드 미확인"

    sido, sigungu, dongri = codes
    san, bun1, bun2 = split_jibun(jibun)
    if not bun1:
        return None, "지번 파싱 실패"

    params = {
        "search_detail_gbn": "2",
        "notice_year": "",
        "notice_year_nm": "",
        "sido": sido["CODE"],
        "sido_nm": sido["NAME"],
        "sigungu": sigungu["CODE"],
        "sigungu_nm": sigungu["NAME"],
        "road_reg": sigungu["CODE"],
        "road_initial": "",
        "road_initial_nm": "",
        "road_code": "",
        "road_code_nm": "",
        "dongri": dongri["CODE"],
        "dongri_nm": dongri["NAME"],
        "reg": sigungu["CODE"],
        "eub": dongri["CODE"],
        "san": san,
        "bun1": bun1,
        "bun2": bun2,
        "build_bun1": "",
        "build_bun2": "00000",
    }
    rows = get_list("/notice/m/gsi/getList.do", params)
    fetched_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if not rows:
        return None, "조회결과 없음"

    if isinstance(trade_date, datetime):
        trade_year = trade_date.year
    else:
        m = re.match(r"(\d{4})", str(trade_date or ""))
        trade_year = int(m.group(1)) if m else None
    chosen = None
    if trade_year:
        chosen = next((row for row in rows if str(row.get("base_year")) == str(trade_year)), None)
    chosen = chosen or rows[0]
    price = to_number(chosen.get("gakuka_w"))
    return {
        "price": price,
        "base_year": chosen.get("base_year"),
        "notice_ymd": chosen.get("notice_ymd"),
        "address": chosen.get("addr"),
        "rows": len(rows),
        "query": {
            "si_gun_gu": norm(si_gun_gu),
            "dong": norm(dong),
            "jibun": norm(jibun),
            "trade_year": trade_year,
            "sido": sido["NAME"],
            "sigungu": sigungu["NAME"],
            "dongri": dongri["NAME"],
            "san": san,
            "bun1": bun1,
            "bun2": bun2,
        },
        "source_url": "https://www.realtyprice.kr/notice/m/gsi/search.do",
        "api_method": "POST",
        "api_endpoint": "https://www.realtyprice.kr/notice/m/gsi/getList.do",
        "request_params": json.dumps(params, ensure_ascii=False, sort_keys=True),
        "verify_url": f"{public_base_url()}/verify.html?{urllib.parse.urlencode(params)}",
        "fetched_at": fetched_at,
        "freshness": "실시간 최신 조회",
    }, "조회완료"


def crawl_columns(crawled, status):
    if not crawled:
        return [status]
    query = crawled.get("query", {})
    price = crawled.get("price")
    price_text = f"{price:,}" if isinstance(price, (int, float)) else norm(price)
    return [
        f"실시간 최신 조회({crawled.get('fetched_at', '')}) {status}: "
        f"{query.get('si_gun_gu', '')} {query.get('dong', '')} {query.get('jibun', '')}, "
        f"{crawled.get('base_year', '')}년 {price_text}, 결과 {crawled.get('rows', '')}건"
    ]


def abbreviation(zone):
    text = norm(zone)
    mapping = {
        # 주거지역
        "제1종전용주거지역": "1전주",
        "제2종전용주거지역": "2전주",
        "제1종일반주거지역": "1주",
        "제2종일반주거지역": "2주",
        "제3종일반주거지역": "3주",
        "준주거지역": "준주거",
        # 상업지역
        "중심상업지역": "중상",
        "일반상업지역": "상업",
        "근린상업지역": "근상",
        "유통상업지역": "유통",
        # 공업지역
        "전용공업지역": "전공",
        "일반공업지역": "일공",
        "준공업지역": "준공",
        # 녹지지역
        "보전녹지지역": "보녹",
        "생산녹지지역": "생녹",
        "자연녹지지역": "자녹",
        # 관리지역
        "보전관리지역": "보관",
        "생산관리지역": "생관",
        "계획관리지역": "계관",
        # 기타
        "농림지역": "농림",
        "자연환경보전지역": "자보",
    }
    return mapping.get(text, text)


def _crawl_task(task):
    idx, total, si_gun_gu, dong, jibun, trade_date, has_price = task
    verb = "검증 조회" if has_price else "조회"
    log(f"{idx}/{total} {norm(dong)} {norm(jibun)} 공시지가 {verb} 중입니다.")
    try:
        crawled, status = crawl_price(si_gun_gu, dong, jibun, trade_date)
        return idx, crawled, status
    except Exception as exc:
        prefix = "원본 유지, " if has_price else ""
        return idx, None, f"{prefix}조회실패: {exc}"


def find_land_row(ws, row):
    if norm(ws.cell(row, 6).value) == "토지":
        return row
    jibun = norm(ws.cell(row, 5).value)
    for candidate in range(row + 1, min(ws.max_row, row + 5) + 1):
        if norm(ws.cell(candidate, 6).value) == "토지":
            if not jibun or norm(ws.cell(candidate, 5).value) == jibun:
                return candidate
            return candidate
    return row


def selected_rows(ws):
    rows = []
    for row in range(5, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if isinstance(value, int):
            rows.append(row)
    return rows


def classify_review(ws, row):
    detail = norm(ws.cell(row, 6).value)
    memo = " ".join(norm(ws.cell(row, col).value) for col in (15, 18))
    if detail == "토지" or "토지만" in memo:
        return "토지만"
    if "철거" in memo:
        return "토지만"
    if "화체" in memo:
        return "토지건물"
    return "토지건물" if detail in ("건물", "") else ""


def copy_widths(src, dst):
    widths = [
        8, 24, 14, 12, 10, 8, 12, 12, 12, 10, 9, 16, 14, 14, 10, 10, 12, 10,
        38,
    ]
    for idx, width in enumerate(widths, start=1):
        dst.column_dimensions[get_column_letter(idx)].width = width


def strip_excel_repair_triggers(wb):
    """Remove legacy drawing/external-link parts that Excel repairs after openpyxl save."""
    if hasattr(wb, "_external_links"):
        wb._external_links = []
    for sheet in wb.worksheets:
        if hasattr(sheet, "_images"):
            sheet._images = []
        if hasattr(sheet, "_charts"):
            sheet._charts = []
        if hasattr(sheet, "legacy_drawing"):
            sheet.legacy_drawing = None


def clone_visible_workbook(wb_values):
    """Create a clean workbook from visible values/styles, leaving broken XML behind."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for src in wb_values.worksheets:
        if src.title == "자동정리":
            continue
        dst = wb.create_sheet(src.title)
        dst.freeze_panes = src.freeze_panes
        dst.sheet_view.showGridLines = src.sheet_view.showGridLines
        if src.auto_filter and src.auto_filter.ref:
            dst.auto_filter.ref = src.auto_filter.ref

        for key, dim in src.column_dimensions.items():
            target = dst.column_dimensions[key]
            target.width = dim.width
            target.hidden = dim.hidden
            target.bestFit = dim.bestFit

        for idx, dim in src.row_dimensions.items():
            target = dst.row_dimensions[idx]
            target.height = dim.height
            target.hidden = dim.hidden

        for row in src.iter_rows():
            for cell in row:
                dst_cell = dst[cell.coordinate]
                dst_cell.value = cell.value
                dst_cell.font = copy(DEFAULT_FONT)
                if cell.has_style:
                    dst_cell.fill = copy(cell.fill)
                    dst_cell.border = copy(cell.border)
                    dst_cell.alignment = copy(cell.alignment)
                    dst_cell.number_format = cell.number_format
                    dst_cell.protection = copy(cell.protection)
                if cell.comment:
                    dst_cell.comment = copy(cell.comment)
                if cell.hyperlink:
                    dst_cell.hyperlink = copy(cell.hyperlink)

        for merged_range in src.merged_cells.ranges:
            dst.merge_cells(str(merged_range))

    return wb


def build(input_path, output_path, summary_path):
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    Path(summary_path).parent.mkdir(parents=True, exist_ok=True)
    log("원본 엑셀을 읽는 중입니다.")
    wb_values = openpyxl.load_workbook(input_path, data_only=True)
    src_values = wb_values.worksheets[0]
    wb = clone_visible_workbook(wb_values)
    strip_excel_repair_triggers(wb)
    if "자동정리" in wb.sheetnames:
        del wb["자동정리"]
    ws = wb.create_sheet("자동정리")

    head_fill = PatternFill("solid", fgColor="D9E8E6")
    thin = Side(style="thin", color="C9D4D1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(2, col, header)
        cell.fill = head_fill
        cell.font = copy(HEADER_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    copy_widths(src_values, ws)
    ws.freeze_panes = "A3"
    ws["A1"] = "주석: 붉은색 텍스트는 realtyprice.kr API에서 캐시 없이 실시간 최신 조회로 가져온 정보입니다."
    ws["A1"].font = copy(CRAWLED_FONT)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

    warnings = []
    rows = selected_rows(src_values)
    log(f"선택 사례 {len(rows)}건을 정리하는 중입니다.")

    # Phase 1: 원본 엑셀에서 모든 행 데이터 추출
    rows_data = []
    for idx, row in enumerate(rows, start=1):
        land_row = find_land_row(src_values, row)
        official_price_orig = src_values.cell(row, 26).value
        rows_data.append({
            "idx": idx,
            "row": row,
            "source_no": src_values.cell(row, 1).value,
            "si_gun_gu": src_values.cell(row, 3).value,
            "dong": src_values.cell(row, 4).value,
            "jibun": src_values.cell(row, 5).value,
            "trade_date": src_values.cell(row, 9).value,
            "land_area": src_values.cell(land_row, 10).value,
            "official_price_orig": official_price_orig,
            "unit_price": src_values.cell(row, 16).value,
            "public_ratio_orig": src_values.cell(row, 27).value,
            "col7": src_values.cell(row, 7).value,
            "col8": src_values.cell(row, 8).value,
            "col12": src_values.cell(row, 12).value,
            "col25": src_values.cell(row, 25).value,
            "review": classify_review(src_values, row),
            "has_price": official_price_orig not in (None, ""),
        })

    # Phase 2: 병렬 API 크롤링
    tasks = [
        (d["idx"], len(rows_data), d["si_gun_gu"], d["dong"],
         d["jibun"], d["trade_date"], d["has_price"])
        for d in rows_data
    ]
    crawl_results = {}
    with ThreadPoolExecutor(max_workers=MAX_CRAWL_WORKERS) as executor:
        for idx, crawled, status in executor.map(_crawl_task, tasks):
            crawl_results[idx] = (crawled, status)

    # Phase 3: 결과를 엑셀에 기록
    preview = []
    for d in rows_data:
        idx = d["idx"]
        crawled, crawl_status = crawl_results.get(idx, (None, "미조회"))

        official_price = d["official_price_orig"]
        official_price_verified_by_api = False

        if not d["has_price"]:
            if crawled and crawled.get("price"):
                official_price = crawled["price"]
                official_price_verified_by_api = True
                crawl_data = crawl_columns(crawled, crawl_status)
            else:
                crawl_data = crawl_columns(None, crawl_status)
                if crawled is None and "조회실패" in crawl_status:
                    warnings.append(f"{norm(d['dong'])} {norm(d['jibun'])} - {crawl_status}")
        else:
            if crawled and crawled.get("price"):
                official_price_verified_by_api = True
                crawl_data = crawl_columns(crawled, crawl_status)
            else:
                crawl_data = crawl_columns(None, crawl_status)

        unit_price = d["unit_price"]
        public_ratio = d["public_ratio_orig"]
        if public_ratio in (None, "") and unit_price not in (None, "") and official_price:
            try:
                public_ratio = unit_price / official_price
            except Exception:
                public_ratio = None

        si_gun_gu_text = norm(d["si_gun_gu"])
        dong_text = norm(d["dong"])
        sojaegi = f"{si_gun_gu_text} {dong_text}".strip() if si_gun_gu_text else dong_text
        record = [
            d["source_no"],
            sojaegi,
            display_jibun(d["jibun"]),
            d["trade_date"],
            d["land_area"],
            d["col7"],
            abbreviation(d["col8"]),
            "",
            d["col25"],
            "",
            "",
            d["col12"],
            unit_price,
            official_price,
            public_ratio,
            d["review"],
            "",
            "",
            *crawl_data,
        ]
        out_row = idx + 2
        for col, value in enumerate(record, start=1):
            cell = ws.cell(out_row, col, value)
            cell.font = copy(DEFAULT_FONT)
            cell.border = border
            cell.alignment = Alignment(
                vertical="center",
                horizontal="left" if col in (3, 19) else None,
            )
            if isinstance(value, datetime):
                cell.number_format = "yyyy-mm-dd"
            if col == 3 and isinstance(value, str):
                cell.number_format = "@"
            if col in (12, 13, 14):
                cell.number_format = "#,##0"
            if col == 15:
                cell.number_format = "0.0000"
            if crawled and col >= 19:
                cell.font = copy(CRAWLED_FONT)
                if col == 19 and crawled.get("verify_url"):
                    cell.hyperlink = crawled["verify_url"]
                    cell.style = "Hyperlink"
                    cell.font = copy(CRAWLED_FONT)
            if official_price_verified_by_api and col == 14:
                cell.font = copy(CRAWLED_FONT)
        preview.append({HEADERS[i]: record[i] for i in range(len(HEADERS))})

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(HEADERS)):
        for cell in row:
            cell.border = border
    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{ws.max_row}"

    log("엑셀 파일을 저장하는 중입니다.")
    wb.save(output_path)
    summary = {
        "sheetName": "자동정리",
        "caseCount": len(rows),
        "preview": json.loads(json.dumps(preview[:100], default=str, ensure_ascii=False)),
        "warnings": warnings,
    }
    Path(summary_path).write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    build(sys.argv[1], sys.argv[2], sys.argv[3])
